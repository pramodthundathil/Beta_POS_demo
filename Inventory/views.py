
from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from django.contrib import messages
from .forms import ProductForm, InventoryStockForm, PurchaseForm, VendorForm, CustomerForm, BatchForm
from django.http import HttpResponse
from POS.models import *
from django.contrib.auth.decorators import login_required
from Finance.models import Income, Expence
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string



# Create your views here.

# units add updating Functions#####################################################

def list_units(request):
    units = Units.objects.all().order_by("-id")
    return render(request,"list-units.html",{"units":units})
# View for adding a unit

def add_unit(request):
    if request.method == 'POST':
        unit_name = request.POST['unit']
        description = request.POST['description']
        new_unit = Units(unit=unit_name, description=description)
        new_unit.save()
        messages.success(request, 'Unit added successfully!')
        return redirect('list_units')  # Assume you have a unit list page
    return render(request, 'add-units.html')


# View for updating a unit
def update_unit(request, unit_id):
    unit = get_object_or_404(Units, pk=unit_id)
    if request.method == 'POST':
        unit.unit = request.POST['unit']
        unit.description = request.POST['description']
        unit.save()
        messages.success(request, 'Unit updated successfully!')
        return redirect('update_unit', unit_id = unit_id)  # Redirect to unit list after update
    return render(request, 'update-units.html', {'unit': unit})

# deleting units...
def delete_unit(request,pk):
    Units.objects.get(id = pk).delete()
    messages.error(request,"unit deleted")
    return redirect("list_units")
########################################## vendor management ############################################
# 

# add a vendor 
@login_required(login_url='SignIn')
def add_vendor(request):
    if request.method == "POST":
        name = request.POST['name']
        email = request.POST['email']
        phone = request.POST['phone']
        gst_number = request.POST['gst_number']
        city = request.POST['city']
        state = request.POST['state']
        country = request.POST['country']
        pincode = request.POST['pincode']
        contact_info = request.POST.get('contact_info', '')
        supply_product = request.POST['supply_product']

        vendor = Vendor(
            name=name,
            email=email,
            phone_number=phone,
            gst_number=gst_number,
            city=city,
            state=state,
            country=country,
            pincode=pincode,
            contact_info=contact_info,
            supply_product=supply_product,
        )
        vendor.save()
        messages.success(request, 'Vendor added successfully')
        return redirect('list_vendor')
    return render(request,"add-vendor.html") 

# displaying vendor list 
@login_required(login_url='SignIn')
def list_vendor(request):
    vendor = Vendor.objects.all().order_by("-id")

    context = {
        "vendor":vendor
    }
    return render(request,'list-vendors.html',context)

#updating vendor
@login_required(login_url='SignIn')
def update_vendor(request, pk):
    vendor = get_object_or_404(Vendor, pk=pk)
    if request.method == 'POST':
        form = VendorForm(request.POST, instance=vendor)
        if form.is_valid():
            form.save()
            return redirect('list_vendor')  # Redirect to a list or relevant page
    else:
        form = VendorForm(instance=vendor)
    return render(request, 'vendor_update.html', {'form': form})

# deleting vendor 
@login_required(login_url='SignIn')
def delete_vendor(request,pk):
    vendor = Vendor.objects.get(id = pk)
    vendor.delete()
    messages.info(request,"vendor deleted....")
    return redirect("list_vendor")


############################ Inventory Management #################################


#add inventory it can be added new inventory in inventory list

@login_required(login_url='SignIn')
def add_inventory(request):
    if request.method == "POST":
        form = InventoryStockForm(request.POST)
        if form.is_valid():
            inventory = form.save()
            inventory.save()  # Save the inventory stock to the database
            messages.success(request, 'Inventory added successfully')
            return redirect('list_inventory')
        else:
            # If form is invalid, render the same page with error messages
            messages.error(request, 'Error adding inventory. Please check the form and try again.')
            return render(request, 'inventory/add-inventory.html', {'forms': form})
    else:
        form = InventoryStockForm()
    return render(request, 'inventory/add-inventory.html', {'forms': form})



@login_required(login_url='SignIn')
def edit_inventory(request,pk):
    inventory  = get_object_or_404(InventoryStock,id = pk)
    purchase = Purchase.objects.filter(purchase_item = inventory)
    form = InventoryStockForm(instance=inventory)
    if request.method == "POST":
        form = InventoryStockForm(request.POST,instance=inventory)
        if form.is_valid():
            inventory = form.save()
            if inventory.product_stock <= inventory.min_stock_level:
                inventory.stock_alert = True
                if Notification.objects.filter(ref_number = inventory.product_code).exists():
                    notification = Notification.objects.filter(ref_number = inventory.product_code).last()
                    notification.updated_at = timezone.now()
                    notification.message = f"Low Stock On Inventory {inventory.product_name} stock level {inventory.product_stock} "
                    notification.save()
                else:
                    notification = Notification(
                        notification_heading = f"Low Stock On {inventory.product_name}",
                        message = f"Low Stock On Inventory {inventory.product_name} stock below {inventory.min_stock_level} ",
                        ref_number = inventory.product_code
                    )
                    notification.save()
        
            
            messages.success(request, 'Inventory Updated successfully')
            return redirect('list_inventory')  # Adjust this based on your URLs
        else:
            messages.error(request, 'Failed to add Purchase Order. Please check the details.')
    
    context = {
        'form': form,
        "inventory":inventory,
        "purchase":purchase
    }
    return render(request, 'inventory/update-inventory.html', context)




@login_required(login_url='SignIn')
def list_inventory(request):
    product = InventoryStock.objects.all().order_by("-id")
    context = {
        "product":product
    }
    return render(request,'inventory/list-inventory.html',context)


@login_required(login_url='SignIn')
def delete_inventory(request,pk):
    inventory = get_object_or_404(InventoryStock,id = pk)
    inventory.delete()
    messages.success(request,"Inventory Deleted successfully....")
    return redirect(list_inventory)


######################################### Purchases and Purchase order ######################################
# Create purchase order for purchase #

@login_required(login_url='SignIn')
def add_purchase_order(request):
    purchase_order = PurchaseOrder.objects.create()
    purchase_order.save()
    return redirect(edit_purchase_order, pk = purchase_order.id)

   
@login_required(login_url='SignIn')
def add_purchase_order_item(request,pk):
    purchase_order = get_object_or_404(PurchaseOrder, id = pk)
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        print(product_id,")))))))))))))))))))))))))))))))))))))))))))")
        try:
            order = PurchaseOrder.objects.get(id=pk)
            if order.save_status == True:
                return JsonResponse({"success": False, "error": "Cannot Be added New Item to This order"})
            product = InventoryStock.objects.get(id=product_id)
            order_item, created = PurchaseOrderItem.objects.get_or_create(purchase_order=order, inventory=product)
            if not created:
                order_item.unit_price = 0
                order_item.quantity += 1
                order_item.save()

            
            # Update order totals
            order.update_totals()
            
            # Render the order items table
            order_items_html = render_to_string('ajaxtemplates/purchase_order_table.html', {'order': order})
            return JsonResponse({"success": True, "html": order_items_html})
        except Order.DoesNotExist:
            return JsonResponse({"success": False, "error": "Order not found"})
        except Product.DoesNotExist:
            return JsonResponse({"success": False, "error": "Product not found"})
    return JsonResponse({"success": False, "error": "Invalid request"})



@login_required(login_url='SignIn')
def list_purchase_order(request):
    order = PurchaseOrder.objects.all().order_by("-id")
    context = {
        "order":order
    }
    return render(request,'inventory/list-purchase-order.html',context)


@login_required(login_url='SignIn')
def edit_purchase_order(request,pk):
    purchase_order  = get_object_or_404(PurchaseOrder,id = pk)
    supplier = Vendor.objects.all()
    product = InventoryStock.objects.all()
    context = {
        "supplier":supplier,
        "product":product,
        "order":purchase_order
    }
    return render(request, 'inventory/add-purchase-order.html', context)


@login_required(login_url='SignIn')
@csrf_exempt
def update_purchase_order_item(request, order_id):
    if request.method == "POST":
        order = get_object_or_404(PurchaseOrder, id=order_id)
        if order.save_status == False:
            item_id = request.POST.get('item_id')
            unit_price = float(request.POST.get('unit_price', 0))
            discount = 0
            print(unit_price,"--------------------")
            quantity = int(request.POST.get('quantity', 1))
             

            # Find the OrderItem to update
            order_item = get_object_or_404(PurchaseOrderItem, id=item_id, purchase_order=order)

            # Update the OrderItem fields
            order_item.unit_price = unit_price
            order_item.discount = discount
            order_item.quantity = quantity
            order_item.save()  # This will also update the total_price based on save() logic
        else:
            return JsonResponse({"success": False, "error": "Cannot be change order is closed"})

        try:
        # Update order totals
            order.update_totals()
            
            print(order.amount)
        # Prepare the updated data to return as a JSON response
            return JsonResponse({
                'success': True,
                'total_amount': order.amount,
                
            })
        except Order.DoesNotExist:
            return JsonResponse({"success": False, "error": "Order not found"})
        except Product.DoesNotExist:
            return JsonResponse({"success": False, "error": "Product not found"})
    return JsonResponse({"success": False, "error": "Invalid request"})



@login_required(login_url='SignIn')
@csrf_exempt
def update_supplier_to_purchase_order(request):
    if request.method == 'POST':
        customer_id = request.POST.get('customer_id')
        order_id = request.POST.get('order_id')
        try:
            order = PurchaseOrder.objects.get(id=order_id)
            customer = Vendor.objects.get(id=customer_id)
            order.supplier = customer
            order.save()
            customer_details_html = render_to_string('ajaxtemplates/suppierinfo.html', {'customers': customer,"order" : order})
            print(customer_details_html)
            return JsonResponse({"success": True, "html": customer_details_html})
        except Order.DoesNotExist:
            return JsonResponse({"success": False, "error": "Order not found"})
        except Customer.DoesNotExist:
            return JsonResponse({"success": False, "error": "Customer not found"})
    return JsonResponse({"success": False, "error": "Invalid request"})


@login_required(login_url='SignIn')
def change_purchase_order_date(request,pk):
    order = get_object_or_404(PurchaseOrder, id= pk)
    if request.method =="POST":
        date = request.POST.get("date")
        order.bill_date = date
        order.save()
        messages.success(request, 'Order Date Changed')
        return redirect("edit_purchase_order",pk = pk)


@login_required(login_url='SignIn')
def purchase_order_invoice(request,pk):
    purchase_order = PurchaseOrder.objects.get(id = pk)

    return render(request,"purchase_order.html",{"purchase_order":purchase_order})


@login_required(login_url='SignIn')
def purchase_from_order(request, order_id):
    purchase_order = get_object_or_404(PurchaseOrder, id=order_id)
    purchase_order.create_purchase()
    purchase_order.order_status = "Closed"
    purchase_order.save()
    messages.info(request,"Purchase Created....")
    return redirect("purchase")

@login_required(login_url='SignIn')
def delete_purchase_order(request, pk):
    purchase_order = get_object_or_404(PurchaseOrder, id = pk)
    purchase_order.delete()
    messages.success(request,"Purchase order deleted success....")
    return redirect("list_purchase_order")

@login_required(login_url='SignIn')
@csrf_exempt
def  delete_purchase_order_item(request,pk):
    if request.method == 'POST':
        itemid = request.POST.get('item_id')
        try:
            order = PurchaseOrder.objects.get(id=pk)
            if order.save_status == True:
                return JsonResponse({"success": False, "error": "Cannot delete item from this order . Closed order"})
            else:
                item = get_object_or_404(PurchaseOrderItem, id = int(itemid))
                item.delete()
            # Update order totals
                order.update_totals()
             
            # Render the order items table
            order_items_html = render_to_string('ajaxtemplates/purchase_order_table.html', {'order': order})
            return JsonResponse({"success": True, "html": order_items_html})
        except Order.DoesNotExist:
            return JsonResponse({"success": False, "error": "Order not found"})
        except Product.DoesNotExist:
            return JsonResponse({"success": False, "error": "Product not found"})
    return JsonResponse({"success": False, "error": "Invalid request"})



#################### Purchase order section ends #######################################################

@login_required(login_url='SignIn')
def purchase(request):
    purchase = Purchase.objects.all().order_by("-id")
    context = {
       "purchase":purchase 
    }
    return render(request,'purchase.html',context)
import datetime

@login_required(login_url='SignIn')
def edit_purchase(request,pk):
    purchase  = get_object_or_404(Purchase,id = pk)
    supplier = Vendor.objects.all()
    product = InventoryStock.objects.all()
    # paid_amount = purchase.paid_amount
    # form = PurchaseForm(instance=purchase)
    # if request.method == "POST":
    #     form = PurchaseForm(request.POST,instance=purchase)
    #     if form.is_valid():
    #         new_purchase = form.save()
    #         new_purchase.save()
    #         now_paid = new_purchase.paid_amount - paid_amount
    #         print(now_paid,"-----------------------------")
    #         if now_paid > 0:
    #             expence = Expence(
    #                 perticulers = f"Amount Paid to  {new_purchase.supplier} towards purchase {new_purchase.purchase_bill_number}",
    #                 date = datetime.datetime.now(),
    #                 bill_number = new_purchase.purchase_bill_number,
    #                 amount = now_paid,
    #                 other = new_purchase.supplier.name if new_purchase.supplier else 'No Partner'
    #             )
    #             expence.save()

    #         messages.success(request, 'Purchase Updated successfully')
    #         return redirect('purchase')  # Adjust this based on your URLs
    #     else:
    #         messages.error(request, 'Failed to add Purchase Order. Please check the details.')
    
    context = {
        # 'form': form,
        "order":purchase,
        "supplier":supplier,
        "product":product
    }
    return render(request, 'inventory/update-purchase.html', context)


@csrf_exempt
def payment_given_in_expense_purchase(request):
    if request.method == "POST":
        order = Purchase.objects.get(id = int(request.POST.get("order_id")))
        if order.purchase_confirmation == False:
            for i in order.purchase_bill.all():
                i.inventory.product_stock += i.quantity
                i.inventory.last_purchase_date = order.bill_date
                i.inventory.last_purchase_amount += i.unit_price
                i.inventory.save()
            order.purchase_confirmation = True
            order.save()
        try:
            if Expence.objects.filter(bill_number = order.purchase_bill_number).exists():
                expense = Expence.objects.filter(bill_number = order.purchase_bill_number)
                total = 0
                
                for ex in expense:
                    total = total + ex.amount
                
                amount = order.paid_amount - total
                if amount > 0:
                    expense = Expence(
                        perticulers = f"Amount paid Against purchase {order.purchase_bill_number}",
                        amount =  round(amount, 2),
                        bill_number = order.purchase_bill_number,
                        other = order.supplier.name if order.supplier else 'No Partner'
                    
                    )
                
                    expense.save() 
                    return JsonResponse({"success": True, "mssg": "Order Payment Updated"})
                else:
                    return JsonResponse({"success": True, "mssg": "Order Payment Updated"})

            else:
                if order.paid_amount > 0:
                    expense = Expence(
                            perticulers = f"Amount paid Against Purchase {order.purchase_bill_number}",
                            amount =  order.paid_amount,
                            bill_number = order.purchase_bill_number,
                            other = order.supplier.name if order.supplier else 'No Partner'
                        
                        )
                    
                    expense.save()
                    return JsonResponse({"success": True, "mssg": "Order Payment Updated"})
                
        except Order.DoesNotExist:
            return JsonResponse({"success": False, "error": "Order not found"})
     
    return JsonResponse({"success": False, "error": "Invalid request"})

          
    

@login_required(login_url='SignIn')
def add_purchase_item(request,pk):
    purchase_order = get_object_or_404(Purchase, id = pk)
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        print(product_id,")))))))))))))))))))))))))))))))))))))))))))")
        try:
            order = Purchase.objects.get(id=pk)
            if order.purchase_confirmation == True:
                return JsonResponse({"success": False, "error": "Cannot Be added New Item to This order"})
            product = InventoryStock.objects.get(id=product_id)
            order_item, created = PurchaseItems.objects.get_or_create(purchase=order, inventory=product)
            if not created:
                order_item.unit_price = 0
                order_item.quantity += 1
                order_item.save()

            
            # Update order totals
            order.update_totals()
            
            # Render the order items table
            order_items_html = render_to_string('ajaxtemplates/purchase_table.html', {'order': order,'total_amount': order.amount})
            return JsonResponse({"success": True, "html": order_items_html})
        except Order.DoesNotExist:
            return JsonResponse({"success": False, "error": "Order not found"})
        except Product.DoesNotExist:
            return JsonResponse({"success": False, "error": "Product not found"})
    return JsonResponse({"success": False, "error": "Invalid request"})

@login_required(login_url='SignIn')
@csrf_exempt
def update_purchase_item(request, order_id):
    if request.method == "POST":
        order = get_object_or_404(Purchase, id=order_id)
        if order.purchase_confirmation == False:
            item_id = request.POST.get('item_id')
            unit_price = float(request.POST.get('unit_price', 0))
            # discount = 0
            print(unit_price,"--------------------")
            quantity = int(request.POST.get('quantity', 1))
             

            # Find the OrderItem to update
            order_item = get_object_or_404(PurchaseItems, id=item_id, purchase=order)

            # Update the OrderItem fields
            order_item.unit_price = unit_price
            # order_item.discount = discount
            order_item.quantity = quantity
            order_item.save()
        else:
            return JsonResponse({"success": False, "error": "This Purchase is already done cannot be Edited "})

              # This will also update the total_price based on save() logic
        try:
        # Update order totals
            order.update_totals()
            
            print(order.amount)
            # order_items_html = render_to_string('ajaxtemplates/purchase_table.html', {'order': order,'total_amount': order.amount})
            # return JsonResponse({"success": True, "html": order_items_html})
        # Prepare the updated data to return as a JSON response
            return JsonResponse({
                'success': True,
                'total_amount': order.amount,
                "balance_amount":order.balance_amount,
                "payment-status":order.payment_status
                
            })
        except Order.DoesNotExist:
            return JsonResponse({"success": False, "error": "Order not found"})
        except Product.DoesNotExist:
            return JsonResponse({"success": False, "error": "Product not found"})
    return JsonResponse({"success": False, "error": "Invalid request"})


@csrf_exempt
def update_purchase_payment(request, order_id):
    if request.method == 'POST':
        payed_amount = float(request.POST.get('payed_amount'))
        # discount = float(request.POST.get('discount'))
        
        try:
            order = Purchase.objects.get(id=order_id)    
            order.paid_amount = payed_amount
            # order.discount = discount
            # order.balance_amount = (order.amount - payed_amount) - discount

            
                        
            if payed_amount == 0:
                order.payment_status = 'UNPAID'
            elif payed_amount >= order.amount:
                order.payment_status = 'PAID'
            else:
                order.payment_status = 'PARTIALLY'
                
            order.save()
            order_items_html = render_to_string('ajaxtemplates/purchase_table.html', {'order': order})
            return JsonResponse({"success": True, "html": order_items_html})
            # return JsonResponse({'success': True})
        except Order.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Order not found'})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required(login_url='SignIn')
@csrf_exempt
def  delete_purchase_item(request,pk):
    if request.method == 'POST':
        itemid = request.POST.get('item_id')
        try:
            order = Purchase.objects.get(id=pk)
            if order.purchase_confirmation == True:
                return JsonResponse({"success": False, "error": "Cannot Be Deleted  Item from This Purchase"})
            else:
                item = get_object_or_404(PurchaseItems, id = int(itemid))
                item.delete()
            # Update order totals
                order.update_totals()
             
            # Render the order items table
            order_items_html = render_to_string('ajaxtemplates/purchase_table.html', {'order': order})
            return JsonResponse({"success": True, "html": order_items_html})
        except Order.DoesNotExist:
            return JsonResponse({"success": False, "error": "Order not found"})
        except Product.DoesNotExist:
            return JsonResponse({"success": False, "error": "Product not found"})
    return JsonResponse({"success": False, "error": "Invalid request"})


@login_required(login_url='SignIn')
@csrf_exempt
def add_bill_discount_to_purchase(request,pk):
    order = Purchase.objects.get(id = pk)
    if request.method == "POST":
        discount = request.POST["bill_discount"]
        order.discount = discount
        order.save()
        order.update_totals()
        # order.calculate_balance()
        messages.success(request,"Discount Added.....")
        return redirect("edit_purchase",pk = pk)

@login_required(login_url='SignIn')
def change_purchase_date(request,pk):
    order = get_object_or_404(Purchase, id= pk)
    if request.method =="POST":
        date = request.POST.get("date")
        order.bill_date = date
        order.save()
        messages.success(request, 'Order Date Changed')
        return redirect("edit_purchase",pk = pk)
    


@login_required(login_url='SignIn')
@csrf_exempt
def update_supplier_to_purchase(request):
    if request.method == 'POST':
        customer_id = request.POST.get('customer_id')
        order_id = request.POST.get('order_id')
        try:
            order = Purchase.objects.get(id=order_id)
            customer = Vendor.objects.get(id=customer_id)
            order.supplier = customer
            order.save()
            customer_details_html = render_to_string('ajaxtemplates/suppierinfo.html', {'customers': customer,"order" : order})
            print(customer_details_html)
            return JsonResponse({"success": True, "html": customer_details_html})
        except Order.DoesNotExist:
            return JsonResponse({"success": False, "error": "Order not found"})
        except Customer.DoesNotExist:
            return JsonResponse({"success": False, "error": "Customer not found"})
    return JsonResponse({"success": False, "error": "Invalid request"})


@login_required(login_url='SignIn')
def deletepurchase(request,pk):
    purchase  = get_object_or_404(Purchase,id = pk)
    purchase.delete()
    messages.error(request, 'Purchase Deleted.....')
    return redirect("purchase")


# @login_required(login_url='SignIn')
# def add_purchase(request):
#     if request.method == 'POST':
#         form = PurchaseForm(request.POST)
#         if form.is_valid():
#             item = form.save()
#             item.save()
#             stock = item.purchase_item
#             if item.paid_amount > 0:
#                 expence = Expence(
#                     perticulers = f"Amount Paid to  {item.supplier} towards purchase {item.purchase_bill_number}",
#                     date = datetime.datetime.now(),
#                     bill_number = item.purchase_bill_number,
#                     amount = item.paid_amount,
#                     other = item.supplier.name if item.supplier else 'No Partner'
#                 )
#                 expence.save()

#             if not stock:
#                 raise ValueError("No inventory item selected for purchase.")

#             # Adjust quantity based on the units in PurchaseOrder and InventoryStock
#             purchase_quantity = item.quantity

#             # If stock is in grams but purchase is in kilograms, convert purchase to grams
#             if stock.unit == 'g' and item.unit == 'kg':
#                 purchase_quantity *= 1000  # Convert kilograms to grams

#             # If stock is in kilograms but purchase is in grams, convert purchase to kilograms
#             elif stock.unit == 'kg' and item.unit == 'g':
#                 purchase_quantity /= 1000  # Convert grams to kilograms

#             stock.product_stock += purchase_quantity
#             stock.last_purchase_date = item.bill_date
#             stock.last_purchase_amount = item.amount
#             stock.save()

#             return redirect('purchase')  
#     else:
#         form = PurchaseForm()
#     return render(request,'inventory/add-purchase.html',{"form":form})

@login_required(login_url='SignIn')
def add_purchase(request):
    if request.method == "POST":
        purchase_type = request.POST.get("purchase_type")
        purchase = Purchase.objects.create(purchase_type = purchase_type)
        purchase.save()
        return redirect(edit_purchase, pk = purchase.id)
    else:
        messages.info(request,"Purchase not created")
        return redirect("purchase")



############################ Product Management #################################


@login_required(login_url='SignIn')
def add_category(request):
    if request.method == "POST":
        name = request.POST['name']
        active = 'active' in request.POST

        category = ProductCategory(
            name=name,
            active=active,
        )
        category.save()
        messages.success(request, 'Category added successfully')
        return redirect('list_category')

    return render(request, 'add-category.html')


@login_required(login_url='SignIn')
def list_category(request):
    category = ProductCategory.objects.all().order_by("-id")
    context = {
        "category":category
    }
    return render(request,"list-category.html",context)

@login_required(login_url='SignIn')
def delete_category(request,pk):
    cat =get_object_or_404(ProductCategory,id = pk)
    cat.delete()
    messages.success(request,"Category Deleted...")
    return redirect("list_category")



@login_required(login_url='SignIn')
def list_products(request):
    product = Product.objects.all().order_by("-id")
    context = {
        "product":product
    }
    return render(request,'list-product.html',context)


@login_required(login_url='SignIn')
def add_product(request):
    form = ProductForm()
    
    if request.method == 'POST':
        form = ProductForm(request.POST)
        ex_date = request.POST.get("ex_date")
        man_date = request.POST.get("man_date")
        if form.is_valid():
            product = form.save()
            product.save()
            try:
                batch = Batch(
                    product = product,
                    expiry_date = ex_date,
                    manufactured_date = man_date,
                    stock_quantity = product.Number_of_stock
                )
                batch.save()
            except:
                batch = Batch(
                    product = product,
                    stock_quantity = product.Number_of_stock
                )
                batch.save()
            inventory = product.inventory  # Get the linked inventory
            try:
            # Fetch the number of units added from the form
                units_to_add = product.Number_of_stock

                # Calculate the total increase in stock based on product's unit_quantity
                total_increase = units_to_add * product.unit_quantity
                print(total_increase,"------------------------------------------------")  # For example, 100g * 10 = 1000g
                
                # Convert the total increase to kilograms if inventory is in kg
                if inventory.unit == 'kg':
                    total_increase_kg = total_increase / 1000  # Convert grams to kilograms
                    inventory.reduce_stock(total_increase_kg)  # Reduce inventory stock by this amount
                else:
                    # If the inventory unit is in grams, reduce directly
                    inventory.reduce_stock(total_increase)

                inventory.save()
        
            except (ValueError, KeyError):
                messages.error(request, "Invalid input. Please enter a valid stock number.")

            return redirect('list_products')  # Redirect to the product list or another page
    
    context = {
        "form":form,

    }
    return render(request,'add-product.html',context)

@login_required(login_url='SignIn')
def product_update(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    food_category = ProductCategory.objects.all()
    tax = Tax.objects.all()
    form = ProductForm(instance=Product)
    batch_form = BatchForm(initial={'product': product})
    batch = Batch.objects.filter(product = product)


    if request.method == 'POST':
        product.name = request.POST['name']
        product.category_id = request.POST['category']
        product.barcode_number = request.POST['barcode']
        product.unit_price = request.POST['price']
        product.Number_of_stock = request.POST['stock']
        product.tax = request.POST['tax']
        product.tax_value_id = request.POST['tax_value']
        product.description = request.POST['description']

        # Handle image upload if a new image is provided
        if 'image' in request.FILES:
            product.image = request.FILES['image']

        # Save updated product
        product.save()

        messages.success(request, 'Product updated successfully!')
        return redirect('product_update', product_id = product_id)  # Redirect to product list or any desired page

    context = {
        'product': product,
        'food_category': food_category,
        'tax': tax,
        "batch_form":batch_form,
        "batch":batch
    }
    return render(request, 'update-product.html', context)


def create_batch(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        form = BatchForm(request.POST)
        if form.is_valid():
            batch = form.save(commit=False)
            batch.product = product  # Assign the product to the batch
            batch.save()

            product.Number_of_stock += batch.stock_quantity
            product.save()
            messages.success(request,"Batch Created.....")
            return redirect('product_update', product_id=product.id)
    else:
        
        return redirect('product_update', product_id=product.id)
    

def update_batch(request,pk):
    batch = get_object_or_404(Batch, id = pk)
    form = BatchForm(instance=batch)
    context = {
        "form":form,
        "batch":batch
    }
    return render(request,"update-batch.html",context)



@login_required(login_url='SignIn')
def incresse_product_stock(request, product_id):
    if request.method == 'POST':
        # Get the product and inventory
        product = get_object_or_404(Product, id=product_id)
        inventory = product.inventory  # Get the linked inventory
        
        try:
            # Fetch the number of units added from the form
            units_to_add = float(request.POST['stock'])
            try:
                batch = get_object_or_404(Batch,id = int(request.POST['batch']))
            except:
                batch = None

            # Calculate the total increase in stock based on product's unit_quantity
            total_increase = units_to_add * product.unit_quantity  # For example, 100g * 10 = 1000g
            
            # Convert the total increase to kilograms if inventory is in kg
            try:
                if inventory.unit == 'kg':
                    total_increase_kg = total_increase / 1000  # Convert grams to kilograms
                    inventory.reduce_stock(total_increase_kg)
                    inventory.save() # Reduce inventory stock by this amount
                else:
                    # If the inventory unit is in grams, reduce directly
                    inventory.reduce_stock(total_increase)
                    inventory.save()

            except:
                messages.info(request,"No Inventory on this Product To Adjust")
            
            # Increase the product stock
            product.Number_of_stock += int(units_to_add)
            try:
                batch.stock_quantity += int(units_to_add)
                batch.save()
            except:
                pass
            
            product.save()
            
            messages.success(request, f"Successfully increased stock for {product.name} by {units_to_add} units.")
        except (ValueError, KeyError):
            messages.error(request, "Invalid input. Please enter a valid stock number.")

    return redirect('product_update', product_id=product.id)


@login_required(login_url='SignIn')
def decrease_product_stock(request, product_id):
    if request.method == 'POST':
        # Get the product and inventory
        product = get_object_or_404(Product, id=product_id)
        inventory = product.inventory  # Get the linked inventory
        
        try:
            # Fetch the number of units added from the form
            units_to_decrease = float(request.POST['stock'])
            try:
                batch = get_object_or_404(Batch,id = int(request.POST['batch']))
            except:
                batch = None

            # Calculate the total increase in stock based on product's unit_quantity
            total_decrease = units_to_decrease * product.unit_quantity  # For example, 100g * 10 = 1000g
            
            # Convert the total increase to kilograms if inventory is in kg
            try:
                if inventory.unit == 'kg':
                    total_increase_kg = total_decrease / 1000  # Convert grams to kilograms
                    # inventory.reduce_stock(total_increase_kg)
                    # inventory.save() # Reduce inventory stock by this amount
                else:
                    # # If the inventory unit is in grams, reduce directly
                    # inventory.reduce_stock(total_increase)
                    # inventory.save()
                    total_increase_kg = total_decrease   # Convert grams to kilograms


            except:
                messages.info(request,"No Inventory on this Product To Adjust")
            
            # Increase the product stock
            product.Number_of_stock -= int(units_to_decrease)
            try:
                batch.stock_quantity -= int(units_to_decrease)
                batch.save()
            except:
                pass
            
            product.save()
            
            messages.success(request, f"Successfully Decreased stock for {product.name} by {units_to_decrease} units.")
        except (ValueError, KeyError):
            messages.error(request, "Invalid input. Please enter a valid stock number.")

    return redirect('product_update', product_id=product.id)


@login_required(login_url='SignIn')
def delete_product(request,pk):
    product = get_object_or_404(Product, id=pk)
    try:
        product.image.delete()
    except:
        pass
    product.delete()
    messages.success(request,"Product Deleted")
    return redirect("list_products")



@login_required(login_url='SignIn')
def disable_product(request,pk):
    try:
        product = Product.objects.get(id = pk)
        if product.status == True:
            product.status = False
        else:
            product.status = True
        product.save()
    except:
        messages.info(request,"Product is not accessible")
    return redirect("list_products")


@login_required(login_url='SignIn')
def AddTax(request):
    if request.method == "POST":
        name = request.POST.get('name')
        tax_rate = request.POST.get('tax')
        tax = Tax.objects.create(tax_name = name,tax_percentage = tax_rate )
        tax.save()
        messages.success(request,'Tax Value Added Success')
        return redirect("ListTax")
    return render(request,"add-tax-slab.html")

@login_required(login_url='SignIn')
def ListTax(request):
    tax = Tax.objects.all()
    context = {
        "tax":tax
    }
    return render(request,"list-tax.html",context)




@login_required(login_url='SignIn')
def add_customer(request,pk):
    if request.method == "POST":
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        email = request.POST.get('email')
        gst = request.POST.get('gst')
        city = request.POST.get('city')
        state = request.POST.get('state')
        country = request.POST.get('country')
        pincode = request.POST.get('pincode')
        contact_info = request.POST.get('contact_info')

        # Creating a new Customer object
        customer = Customer(
            name=name,
            phone=phone,
            email=email,
            gst_number=gst,
            city=city,
            state=state,
            country=country,
            pincode=pincode,
            contact_info=contact_info
        )

        try:
            # Saving the object to the database
            customer.save()
            messages.info(request, "customer addedd.............")
            order = Order.objects.get(id = pk)
            order.customer = customer
            order.save()
            return redirect("POS",pk = pk)

        # Redirect to a success page or the customer's detail page
        except Exception as e:
            return HttpResponse(f"Error: {e}")

    return redirect("POS")


@login_required(login_url='SignIn')
def list_customer(request):
    customer = Customer.objects.all()
    context = {
        "customer":customer
    }
    return render(request,"list-customers.html",context)



@login_required(login_url='SignIn')
def add_customers(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request,"Customer Added.....")
            return redirect('list_customer')  # Redirect to the customer list or relevant page
    else:
        form = CustomerForm()
    return render(request, 'add-customers.html', {'form': form})



# Update customer
@login_required(login_url='SignIn')
def update_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST,request.FILES,instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request,"Customer Updated.....")

            return redirect('list_customer')  # Redirect to the customer list or relevant page
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'update_customer.html', {'form': form,'customer': customer})

# Delete customer
@login_required(login_url='SignIn')
def delete_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    
    customer.delete()
    messages.success(request,"Customer deleted.....")

    return redirect('list_customer')  # Redirect to the customer list after deletion
    


# Purchases.................................................


#Bulk Delete

def delete_bulk_inventory(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            InventoryStock.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("list_inventory")




def delete_bulk_purchase(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            Purchase.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("purchase")


def delete_bulk_purchase_order(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            PurchaseOrder.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("list_purchase_order")


def delete_bulk_products(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            Product.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("list_products")


def delete_bulk_invoice(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            Order.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("list_sale")


def delete_bulk_invoice_pending(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            Order.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("list_sale_pending")


def delete_bulk_invoice_partial(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            Order.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("list_sale_partial")


def delete_bulk_income(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            Income.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("income")

def delete_bulk_expense(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            Expence.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("expence")


def delete_bulk_category(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            ProductCategory.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("list_category")

def delete_bulk_supplier(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            Vendor.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("list_vendor")


def delete_bulk_customers(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            Customer.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("list_customer")


def delete_bulk_staff(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            Staff.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("list_staff")

from Home.models import StaffSalary
def delete_bulk_staff_salary(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('contact_id[]')  # Get the selected IDs from the form
        print(selected_ids,"----------------------------------")
        if selected_ids:
            StaffSalary.objects.filter(id__in=selected_ids).delete()
            messages.success(request, 'Selected items have been deleted.')
        else:
            messages.warning(request, 'No items were selected for deletion.')
    return redirect("list_salary")

import xlrd
from openpyxl import load_workbook

@login_required(login_url="login")
def import_data_from_excel_inventory(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        if str(excel_file).split(".")[-1].lower() == 'xls':
            workbook = xlrd.open_workbook(file_contents=excel_file.read())
            worksheet = workbook.sheet_by_index(0)
        else:
            workbook = load_workbook(excel_file)
            worksheet = workbook.active

        for row in worksheet.iter_rows(min_row=2, values_only=True):

            if row[0] == None or row[1] == None:
                continue
            else:

                print(str(row[0]),str(row[1]),str(row[3]),str(row[3]))
                
                try:
                    product_name = str(row[0])
                except:
                    continue
                try:
                    pro_stock = float(row[1])
                except:
                    continue
                try:
                    unit = str(row[2])
                except:
                    unit = "kg" 
                try:
                    min_stock=float(row[3])
                except:
                    min_stock = 0
      

                if InventoryStock.objects.filter(product_name=product_name,unit = unit ).exists():
                    try:

                        stock = InventoryStock.objects.filter(product_name=product_name,unit = unit )[0]
                        stock.product_stock += pro_stock
                        stock.save()
                    except:
                        continue


                else:
                    InventoryStock.objects.create(
                        product_name = product_name,
                        unit = unit,
                        product_stock = pro_stock,
                        min_stock_level = min_stock
                    )
        messages.info(request,"excel File Updated....")
                
    return redirect("list_inventory")



@login_required(login_url="login")
def import_data_from_excel_product(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        if str(excel_file).split(".")[-1].lower() == 'xls':
            workbook = xlrd.open_workbook(file_contents=excel_file.read())
            worksheet = workbook.sheet_by_index(0)
        else:
            workbook = load_workbook(excel_file)
            worksheet = workbook.active

        for row in worksheet.iter_rows(min_row=2, values_only=True):

            if row[0] == None or row[1] == None:
                continue
            else:

                print(str(row[0]),str(row[1]),str(row[3]),str(row[3]))
                
                try:
                    product_name = str(row[0])
                except:
                    continue
                try:
                    pro_stock = float(row[1])
                except:
                    continue
                try:
                    unit = str(row[2])
                except:
                    unit = "kg" 
                try:
                    min_stock=float(row[3])
                except:
                    min_stock = 0
      

                if InventoryStock.objects.filter(product_name=product_name,unit = unit ).exists():
                    try:

                        stock = InventoryStock.objects.filter(product_name=product_name,unit = unit )[0]
                        stock.product_stock += pro_stock
                        stock.save()
                    except:
                        continue


                else:
                    InventoryStock.objects.create(
                        product_name = product_name,
                        unit = unit,
                        product_stock = pro_stock,
                        min_stock_level = min_stock
                    )
        messages.info(request,"excel File Updated....")
                
    return redirect("list_inventory")



# barcode generation

import barcode
from barcode.writer import ImageWriter
from django.core.files import File
from io import BytesIO

def generate_barcode(batch):
    product  = batch.product
    """Generate a barcode image for the product."""
    if product.barcode_number:
        ean = barcode.get('code128', product.barcode_number, writer=ImageWriter())
        buffer = BytesIO()
        ean.write(buffer)
        return buffer
    return None


def product_barcode_image(request, pk):
    batch = get_object_or_404(Batch,id = pk)
    product = batch.product
    barcode_buffer = generate_barcode(batch = batch)

    if barcode_buffer:
        response = HttpResponse(barcode_buffer.getvalue(), content_type='image/png')
        response['Content-Disposition'] = f'inline; filename={product.barcode_number}.png'
        return response
    else:
        return HttpResponse("No barcode available", status=404)
    

def barcode_view(request,pk):
    batch = get_object_or_404(Batch, id = pk)
    if request.method == "POST":
        num = request.POST.get("bnum")
        return render(request,"barcode.html",{"batch":batch,'range': range(int(num))})
    else:
        return render(request,"barcode.html",{"batch":batch,'range': range(1)})

        



